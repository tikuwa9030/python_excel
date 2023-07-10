from pathlib import Path
from openpyxl import Workbook, load_workbook

# 新規ブック作成
wb_new = Workbook()

# 新規ブックの選択シートを取得
ws_new = wb_new.active

# シート名を入力
ws_new.title = "一覧表"

# 一覧の項目名を入力
ws_new['B2'] = "部署名"
ws_new['C2'] = "氏名"

# 既存ブックの読み込み
path = Path('./PythonExcelサンプルコード/Chapter2/Section010/books')

# ファイルの取得と値を取得して新規ブックに設定
for i, file in enumerate(path.glob('*.xlsx')):

    # 既存ブックの読み込み
    wb = load_workbook(file, read_only=True)

    # シートの取得
    ws = wb['チェックリスト']

    # 値を設定する行番号
    row_no = i + 3

    # 読み込んだファイルの部署名と氏名を転記
    ws_new[f'B{row_no}'] = ws['C2'].value
    ws_new[f'C{row_no}'] = ws['C3'].value

# ファイルを保存
wb_new.save("一覧表.xlsx")
