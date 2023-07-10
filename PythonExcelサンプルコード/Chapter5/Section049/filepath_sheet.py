from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, PatternFill, Side

wb = Workbook()
ws = wb.active
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 30

black_thin = Side(color='000000', border_style='thin')
border = Border(left=black_thin, right=black_thin,
                top=black_thin, bottom=black_thin)
green_fill = PatternFill(fgColor='C6E0B4', fill_type='solid')

for i, col_name in enumerate(('項番', 'パス', 'ブック名')):
    col_no = i + 1
    ws.cell(1, col_no).value = col_name
    ws.cell(1, col_no).fill = green_fill
    ws.cell(1, col_no).border = border


path = Path('./books')
for i, file in enumerate(path.glob('**/*.xlsx')):
    row_no = i + 2
    ws.cell(row_no, 1).value = i + 1

    absolute_path = file.resolve()
    filepath = absolute_path.parent
    filename = absolute_path.name

    ws.cell(row_no, 2).value = str(filepath)
    ws.cell(row_no, 3).value = filename

    ws.cell(row_no, 1).border = border
    ws.cell(row_no, 2).border = border
    ws.cell(row_no, 3).border = border

wb.save('引き継ぎ資料.xlsx')
