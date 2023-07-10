from datetime import date

from openpyxl import load_workbook

wb = load_workbook('申請書.xlsx')
ws = wb.active

ws['C4'].value = '営業部一課'
ws['C5'].value = '佐藤花子'
ws['C6'].value = date.today()

for row in ws.iter_rows(min_row=11, max_row=ws.max_row, min_col=2, max_col=7):
    row[0].value = None
    row[4].value = None
    row[5].value = None

wb.save('申請書_変更後.xlsx')
