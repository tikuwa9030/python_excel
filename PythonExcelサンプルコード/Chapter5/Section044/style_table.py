from openpyxl import load_workbook
from openpyxl.styles import Border, Font, NamedStyle, Side

wb = load_workbook('課題一覧.xlsx')
ws = wb.active

table_style = NamedStyle(name='table_style')
table_style.font = Font(name='Yu Gothic', size=13)
black_thin = Side(color='000000', border_style='thin')
table_style.border = Border(left=black_thin, right=black_thin,
                            top=black_thin, bottom=black_thin)

for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=5):
    for cell in row:
        cell.style = table_style

wb.save('課題一覧_変更後.xlsx')
