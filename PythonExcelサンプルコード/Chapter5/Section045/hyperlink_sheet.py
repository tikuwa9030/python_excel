from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side

wb = load_workbook('業務マニュアル.xlsx')

ws_new = wb.create_sheet(title='目次', index=0)
ws_new.column_dimensions['B'].width = 40
for ws in wb.worksheets:
    ws.sheet_view.tabSelected = None

black_thin = Side(color='000000', border_style='thin')
border = Border(left=black_thin, right=black_thin,
                top=black_thin, bottom=black_thin)
green_fill = PatternFill(fgColor='C6E0B4', fill_type='solid')
hyperlink_style = 'Hyperlink'

for i, col_name in enumerate(('項番', 'ページへのリンク')):
    col_no = i + 1
    ws_new.cell(1, col_no).value = col_name
    ws_new.cell(1, col_no).fill = green_fill
    ws_new.cell(1, col_no).border = border

for i, sheetname in enumerate(wb.sheetnames[1:]):
    row_no = i + 2
    ws_new.cell(row_no, 1).value = i + 1
    ws_new.cell(row_no, 1).border = border

    ws_new.cell(row_no, 2).value = sheetname
    ws_new.cell(row_no, 2).hyperlink = f'#{sheetname}!A1'
    ws_new.cell(row_no, 2).style = hyperlink_style
    ws_new.cell(row_no, 2).border = border

wb.active = 0
wb.save('業務マニュアル_変更後.xlsx')
