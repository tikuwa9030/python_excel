import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

wb = Workbook()
ws = wb.active

df = pd.read_csv('hanbai.csv', encoding='utf-8')
df['販売数'] *= 1000
for row in dataframe_to_rows(df, index=None, header=True):
    ws.append(row)

line = LineChart()
line.style = 13
data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

line.add_data(data, titles_from_data=True)
line.set_categories(labels)

line.x_axis.title = '月'
line.y_axis.title = '販売数'
line.title = '販売数推移'

ws.add_chart(line, 'A12')
wb.save('販売数_折れ線グラフ.xlsx')
