import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows

wb = Workbook()
ws = wb.active

df = pd.read_csv('uriage.csv', encoding='utf-8')
df = df.sort_values(by='当期売上', ascending=False)
for row in dataframe_to_rows(df, index=None, header=True):
    ws.append(row)

pie = PieChart()
pie.style = 37
data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

pie.add_data(data)
pie.set_categories(labels)

pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.title = '部門別売上'

ws.add_chart(pie, 'A9')
wb.save('部門別売上_円グラフ.xlsx')
