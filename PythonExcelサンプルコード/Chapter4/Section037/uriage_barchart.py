import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

wb = Workbook()
ws = wb.active
df = pd.read_csv('uriage.csv', encoding='utf-8')
for row in dataframe_to_rows(df, index=None, header=True):
    ws.append(row)

bar = BarChart()
bar.type = 'col'
data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
bar.add_data(data, titles_from_data=True)
bar.set_categories(labels)
bar.x_axis.title = '部門'
bar.y_axis.title = '売上高（百万円）'
bar.title = '部門別売上高'

ws.add_chart(bar, 'A9')
wb.save('部門別売上高_棒グラフ.xlsx')
