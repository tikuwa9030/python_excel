import jaconv
from openpyxl import load_workbook

wb = load_workbook('チェックリスト_3.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    value = row[0].value
    if value is not None:
        row[0].value = jaconv.h2z(value, digit=True)

wb.save('チェックリスト_3_変更後.xlsx')
