from openpyxl import load_workbook

wb = load_workbook('チェックリスト_2.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    value = row[0].value
    if value is not None:
        row[0].value = value.title()

wb.save('チェックリスト_2_変更後.xlsx')
