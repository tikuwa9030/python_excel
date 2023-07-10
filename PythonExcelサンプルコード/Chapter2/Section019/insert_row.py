from openpyxl import load_workbook

wb = load_workbook('品目別売上.xlsx')
ws = wb.active

num = 3
start_row = 3
for row_no in range(ws.max_row-num+1, num+start_row, -num):
    ws.insert_rows(row_no)

wb.save('品目別売上_変更後.xlsx')
