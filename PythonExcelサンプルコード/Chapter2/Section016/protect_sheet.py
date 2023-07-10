from openpyxl import load_workbook
from openpyxl.styles import Protection

wb = load_workbook('見積書.xlsx')
ws = wb['見積書']

for rows in ws['B11:H24']:
    for cell in rows:
        cell.protection = Protection(locked=False)

ws.protection.password = 'test'
ws.protection.enable()

wb.save('見積書2_変更後.xlsx')
