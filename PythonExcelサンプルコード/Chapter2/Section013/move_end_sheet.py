from openpyxl import load_workbook

wb = load_workbook('チェックリスト.xlsx')

wb.move_sheet('まとめ', offset=len(wb.sheetnames))

wb.save('チェックリスト_変更後.xlsx')
