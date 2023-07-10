from openpyxl import load_workbook

wb = load_workbook('売上実績.xlsx')
ws = wb.active

for row_no in range(2, ws.max_row + 1):
    ws.row_dimensions[row_no].hidden = False

for col_no in range(2, ws.max_column + 1):
    col_alphabet = ws.cell(row=1, column=col_no).column_letter
    ws.column_dimensions[col_alphabet].hidden = False

wb.save('売上実績_変更後.xlsx')
