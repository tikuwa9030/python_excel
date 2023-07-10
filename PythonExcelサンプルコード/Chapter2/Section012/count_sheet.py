from pathlib import Path

from openpyxl import load_workbook, Workbook

wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = '集計'

ws_new['B2'] = 'ブック名'
ws_new['C2'] = '全シート数'
ws_new['D2'] = '非表示シート数'

path = Path('./books')
for i, file in enumerate(path.glob('*.xlsx')):
    wb = load_workbook(file)

    row_no = i + 3
    ws_new[f'B{row_no}'] = file.name
    ws_new[f'C{row_no}'] = len(wb.sheetnames)

    hidden_worksheets = [
        ws for ws in wb.worksheets if ws.sheet_state != ws.SHEETSTATE_VISIBLE]
    ws_new[f'D{row_no}'] = len(hidden_worksheets)

wb_new.save('シート数集計.xlsx')
