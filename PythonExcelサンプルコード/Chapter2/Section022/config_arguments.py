import configparser

from openpyxl import load_workbook

config = configparser.ConfigParser()
config.read('sample.ini', encoding='utf-8')

default = config['DEFAULT']
filename = default['filename']
cellno = default['cellno']

wb = load_workbook(filename, read_only=True)
ws = wb.active

print(ws[cellno].value)
