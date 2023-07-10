import sys

from openpyxl import load_workbook

filename = sys.argv[1]
cellno = sys.argv[2]

wb = load_workbook(filename, read_only=True)
ws = wb.active

print(ws[cellno].value)
