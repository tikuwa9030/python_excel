from pathlib import Path

from openpyxl import load_workbook, Workbook

wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = '台帳'
ws_new.column_dimensions['A'].width = 20

path = Path('./books')
for i, file in enumerate(path.glob('*.xlsx')):
    wb = load_workbook(file, data_only=True)
    ws = wb['請求書']

    row_no = i + 1
    ws_new.cell(row_no, 1).value = ws['B4'].value
    ws_new.cell(row_no, 2).value = ws['H10'].value
    ws_new.cell(row_no, 2).number_format = ws['H10'].number_format

wb_new.save('台帳.xlsx')
