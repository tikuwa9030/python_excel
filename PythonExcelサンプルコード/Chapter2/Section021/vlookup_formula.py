from openpyxl import load_workbook

wb = load_workbook('作業時間表.xlsx')

lastmonth = '202004'
month = '202005'

ws_lastmonth = wb[lastmonth]
ws = wb[month]

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    row_count = row[0].row
    row[4].value = f'= VLOOKUP(B{row_count}, {lastmonth}!$B$2:$D$11, 3, FALSE)'

wb.save('作業時間表_変更後.xlsx')
