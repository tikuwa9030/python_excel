from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

wb = load_workbook('粗利.xlsx')
ws = wb.active

origin_cell_no = 'D6'
ws[origin_cell_no] = '=C6-B6'

for row_no in range(7, ws.max_row + 1):
    cell_no = f'D{row_no}'

    ws[cell_no] = Translator(
        ws[origin_cell_no].value, origin=origin_cell_no).translate_formula(cell_no)

wb.save('粗利_変更後.xlsx')
