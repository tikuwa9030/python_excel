from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

wb = load_workbook('合計.xlsx')
ws = wb.active

ws['D3'] = Translator(ws['C3'].value, origin='C3').translate_formula('D3')

wb.save('合計_変更後.xlsx')
