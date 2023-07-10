import time
from pathlib import Path

from selenium.webdriver import Chrome
from openpyxl import load_workbook

wb = load_workbook('勤怠.xlsx', read_only=True)
ws = wb.active
start_time = ws['B2'].value
end_time = ws['C2'].value

try:
    driver = Chrome()
    html_path = Path('./sample_window.html').resolve()
    driver.get(str(html_path))
    time.sleep(5)

    driver.find_element_by_id('openwindow').click()
    time.sleep(5)

    window_handles = driver.window_handles
    driver.switch_to.window(window_handles[-1])

    start = driver.find_element_by_id('start')
    start.send_keys(str(start_time))

    end = driver.find_element_by_id('end')
    end.send_keys(str(end_time))

    start.submit()
    time.sleep(5)

finally:
    driver.quit()
