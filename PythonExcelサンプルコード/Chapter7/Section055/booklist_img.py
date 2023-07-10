import time
from io import BytesIO
from urllib import parse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image

url = 'https://book.impress.co.jp/'
r = requests.get(url)
soup = BeautifulSoup(r.text, 'html.parser')
books = soup.select('div.block-sub-box-body > ol > li')

wb = Workbook()
ws = wb.active
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 40

for i, book in enumerate(books):
    time.sleep(1)

    row_no = i + 1
    ws.cell(row_no, 1).value = book.text
    ws.cell(row_no, 2).value = parse.urljoin(url, book.find('a')['href'])

    image_url = book.find('img')['src']
    image_r = requests.get(parse.urljoin('http:', image_url))
    image = Image(BytesIO(image_r.content))

    image.width = 80
    image.height = 120
    ws.add_image(image, ws.cell(row_no, 3).coordinate)
    ws.row_dimensions[row_no].height = 100

wb.save('ブックランキング.xlsx')
