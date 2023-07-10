from openpyxl import load_workbook
from openpyxl.styles import Font

wb = load_workbook('売上実績.xlsx')
ws = wb.active

blue_font = Font(name='ＭＳ Ｐ明朝', color='0000FF',
                 size=18, bold=True)

for row in ws['B2':'F2']:
    for cell in row:
        cell.font = blue_font

wb.save('売上実績_変更後.xlsx')
