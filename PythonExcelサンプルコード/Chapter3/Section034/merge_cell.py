from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

wb = load_workbook('作業時間.xlsx')
ws = wb.active
green_fill = PatternFill(fgColor='C6E0B4', fill_type='solid')
center_alignment = Alignment(horizontal='center')
black_thin = Side(color='000000', border_style='thin')
border = Border(left=black_thin, right=black_thin,
                top=black_thin, bottom=black_thin)

row_no = 2
start = 5
end = 16
ws.merge_cells(start_row=row_no, start_column=start,
               end_row=row_no, end_column=end)
ws.cell(row_no, start).fill = green_fill
ws.cell(row_no, start).alignment = center_alignment

for column_no in range(start, end + 1):
    ws.cell(row_no, column_no).border = border

wb.save('作業時間_変更後.xlsx')
