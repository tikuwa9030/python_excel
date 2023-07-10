from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

wb = load_workbook('課題一覧.xlsx')
ws = wb.active

gray_fill = PatternFill(bgColor='C0C0C0', fill_type='solid')
cell_rule = FormulaRule(formula=['$D3="完了"'], fill=gray_fill)
ws.conditional_formatting.add('B3:G12', cell_rule)

wb.save('課題一覧_変更後.xlsx')
