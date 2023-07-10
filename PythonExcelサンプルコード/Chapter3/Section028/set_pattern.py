from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

wb = load_workbook('売上実績.xlsx')
ws = wb.active

green_fill = PatternFill(fgColor='C6E0B4', fill_type='solid')
center_alignment = Alignment(horizontal='center', vertical='center')

for row in ws['B2':'F2']:
    for cell in row:
        cell.fill = green_fill
        cell.alignment = center_alignment

wb.save('売上実績_変更後.xlsx')
