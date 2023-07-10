from openpyxl import load_workbook
from openpyxl.styles import Border, Side

wb = load_workbook('売上実績.xlsx')
ws = wb.active

black_thin = Side(color='000000', border_style='thin')
border = Border(left=black_thin, right=black_thin,
                top=black_thin, bottom=black_thin)

for row in ws.iter_rows(min_row=2, min_col=2):
    for cell in row:
        cell.border = border

wb.save('売上実績_変更後.xlsx')
