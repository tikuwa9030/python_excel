from openpyxl import load_workbook

wb = load_workbook('売上実績.xlsx')
ws = wb.active

for row in ws.iter_rows(min_col=3, max_col=4, min_row=3):
    row[0].number_format = 'yyyy-mm-dd'
    row[1].number_format = '#,##0'

wb.save('売上実績_変更後.xlsx')
