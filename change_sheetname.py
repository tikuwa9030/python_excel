from openpyxl import Workbook, load_workbook


class create:

    def sheets(self):
        """入力した数のシートを作成して、ブックを保存


        """
        # 作成するシート数を入力
        count = input("全シート数は：")

        # ブックを作成
        wb = Workbook()

        # アクティブなシートを選択
        ws = wb.active

        # ワークシートを作成
        ws.title = "概要_1"

        # 入力したシート数分を複製して、シート名をつける
        for i in range(2, int(count) + 1):

            # ワークシートを作成
            wb.create_sheet(title=f"概要_{i}")

        # ブックを保存
        wb.save("資料.xlsx")


class change:

    def sheetname(self):
        """シートを取得し、シート名を変更する


        """
        # ブックのシートを全取得
        wb = load_workbook('資料.xlsx')
        # 取得したシート名とシートカラーを変更
        for i, ws in enumerate(wb.worksheets):
            # タイトル名を変更
            ws.title = "ID_" + ws.title
            # 10シートおきにシートの色を変える
            if (i + 1) % 10 == 0:
                ws.sheet_properties.tabColor = '0000FF'

        # 資料を保存
        wb.save("資料_変更後.xlsx")


if __name__ == '__main__':
    # インスタンス化
    cre = create()
    cha = change()

    # ブック作成
    cre.sheets()

    # シート作成
    cha.sheetname()
